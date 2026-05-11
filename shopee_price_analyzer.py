"""
Shopee Price Analyzer & Competitor Intelligence Tool
---------------------------------------------------
Script ini digunakan untuk melakukan scraping data produk dari Shopee,
membersihkan data dari harga outlier menggunakan metode statistik IQR,
serta memberikan rekomendasi strategi penetapan harga (pricing) untuk Shopee Ads.

Fitur:
- Single & Bulk Keyword Input (.txt)
- Visual DOM Scraping (mengatasi pemblokiran API)
- IQR (Interquartile Range) Outlier Filtering
- Excel (.xlsx) Report Generation
- Auto-detect toko milik sendiri berdasarkan lokasi dan kecocokan harga
"""

import time
import statistics
import urllib.parse
import os
from datetime import datetime

# Optional UI rendering imports
try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich import box
    HAS_RICH = True
except ImportError:
    HAS_RICH = False

# Optional Excel export imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import undetected_chromedriver as uc


def setup_driver() -> uc.Chrome:
    """
    Menginisialisasi dan mengatur undetected_chromedriver.
    Menggunakan profil Chrome lokal agar sesi (seperti login/captcha) tersimpan.
    
    Returns:
        uc.Chrome: Instance dari webdriver Chrome.
    """
    print("\n[INFO] Membuka browser (undetected) untuk Visual Scraping...")
    options = uc.ChromeOptions()
    profile_path = os.path.join(os.getcwd(), "shopee_profile")
    options.add_argument(f"--user-data-dir={profile_path}")
    
    driver = uc.Chrome(options=options, version_main=147)
    driver.set_page_load_timeout(60)   
    driver.set_script_timeout(30)      
    return driver


def scrape_shopee(driver: uc.Chrome, keyword: str, max_items: int = 50, auto_mode: bool = False) -> list:
    """
    Mencari keyword di Shopee dan mengekstrak data produk (nama, harga, terjual, lokasi) dari DOM.
    
    Args:
        driver (uc.Chrome): Instance browser yang sedang berjalan.
        keyword (str): Kata kunci produk yang dicari.
        max_items (int): Maksimal produk yang ingin diekstrak.
        auto_mode (bool): Jika True, tidak akan menunggu input 'Enter' dari user.
        
    Returns:
        list: Daftar dictionary berisi data produk kompetitor.
    """
    products = []
    try:
        safe_keyword = urllib.parse.quote(keyword)
        search_url = f"https://shopee.co.id/search?keyword={safe_keyword}"
        
        print(f"\n[INFO] Membuka link pencarian: {keyword}...")

        try:
            driver.get(search_url)
        except Exception as e:
            print(f"[WARNING] Page load timeout atau lambat: {e}")
        
        if not auto_mode:
            print("\n" + "="*69)
            print(" ⚠️ PERHATIAN: Silakan cek browser Chrome yang baru saja terbuka!")
            print(" Jika muncul halaman 'Select Your Language' / Error / Captcha:")
            print(" 1. Klik 'Bahasa Indonesia' atau selesaikan puzzle secara manual.")
            print("="*69 + "\n")
            input("👉 Jika browser sudah menampilkan produk dengan normal, tekan ENTER di sini...")
        else:
            time.sleep(4) 
        
        print("[INFO] Melakukan scroll perlahan agar semua produk termuat (lazy-loading)...")
        for _ in range(8):
            driver.execute_script("window.scrollBy(0, 700);")
            time.sleep(1.5)
            
        print("[INFO] Membaca data produk, harga, terjual, dan lokasi dari layar...")
        
        # Injeksi JavaScript untuk mengekstrak elemen DOM
        js_script = """
        let results = [];
        let cards = document.querySelectorAll('a[href*="-i."]');
        
        cards.forEach(card => {
            let name = "";
            let img = card.querySelector('img');
            if(img && img.hasAttribute('alt')) {
                name = img.getAttribute('alt');
            } else {
                let divs = card.querySelectorAll('div');
                for(let d of divs) {
                    let txt = d.innerText.trim();
                    if(txt.length > 15 && !txt.includes('Rp') && !txt.includes('Terjual')) {
                        name = txt;
                        break;
                    }
                }
            }
            
            let price = 0; let sold = 0; let lokasi = "-";
            let elements = card.querySelectorAll('span, div');
            
            elements.forEach(el => {
                let txt = el.innerText.trim();
                let txtLower = txt.toLowerCase();
                
                if(txt.startsWith('Rp') && txt.length > 2) {
                    let match = txt.match(/Rp\\s*([\\d\\.]+)/);
                    if(match && price === 0) {
                        let cleanPrice = match[1].replace(/[^0-9]/g, '');
                        price = parseInt(cleanPrice);
                    }
                }
                
                if(txtLower.includes('terjual')) {
                    let matchSold = txtLower.match(/([\\d\\.,]+)\\s*(rb)?\\s*terjual/);
                    if(matchSold && sold === 0) {
                        let numStr = matchSold[1];
                        let isRb = matchSold[2] === 'rb';
                        if(isRb) numStr = numStr.replace(',', '.');
                        else numStr = numStr.replace(/\\./g, '').replace(',', '.');
                        
                        let num = parseFloat(numStr);
                        if(!isNaN(num)) sold = isRb ? Math.round(num * 1000) : Math.round(num);
                    }
                }
            });
            
            let leafTexts = [];
            elements.forEach(el => {
                if(el.children.length === 0 && el.innerText.trim() !== '') leafTexts.push(el.innerText.trim());
            });
            
            for (let i = leafTexts.length - 1; i >= 0; i--) {
                let t = leafTexts[i], tL = t.toLowerCase();
                if (/[a-z]/i.test(tL) && !tL.includes('rp') && !tL.includes('terjual') && 
                    !tL.includes('menit') && !tL.includes('hari') && !tL.includes('bulan') && 
                    !tL.includes('diskon') && !tL.includes('cashback')) {
                    lokasi = t; break; 
                }
            }
            
            if(name && price > 0) {
                results.push({ "nama_produk": name.substring(0, 70).trim(), "harga": price, "total_terjual": sold || 0, "lokasi": lokasi });
            }
        });
        
        return results.filter((v,i,a)=>a.findIndex(t=>(t.nama_produk === v.nama_produk))===i);
        """
        
        try:
            data = driver.execute_script(js_script)
        except Exception as e:
            print(f"\n[WARNING] Script timeout atau error saat membaca DOM: {e}")
            data = None
        
        if data:
            products = data[:max_items]
        else:
            print("\n[WARNING] Tidak ada data terbaca. Pastikan halaman sudah termuat.")
            
    except Exception as e:
        print(f"\n[ERROR SYSTEM] {e}")
            
    return products


def filter_outlier_iqr(competitors: list, multiplier: float = 1.5) -> tuple:
    """
    Menyaring harga yang tidak wajar (outlier) menggunakan Interquartile Range (IQR).
    
    Args:
        competitors (list): Daftar dictionary data kompetitor.
        multiplier (float): Pengali IQR (standar 1.5).
        
    Returns:
        tuple: (list_produk_valid, list_produk_outlier)
    """
    valid_comps = [c for c in competitors if c["harga"] > 0]
    if len(valid_comps) < 4: 
        return valid_comps, []
    
    prices = [c["harga"] for c in valid_comps]
    q1 = statistics.quantiles(prices, n=4)[0]
    q3 = statistics.quantiles(prices, n=4)[2]
    iqr = q3 - q1
    lower = q1 - multiplier * iqr
    upper = q3 + multiplier * iqr
    
    clean_comps = [c for c in valid_comps if lower <= c["harga"] <= upper]
    outlier_comps = [c for c in valid_comps if c["harga"] < lower or c["harga"] > upper]
    
    return clean_comps, outlier_comps


def analisa_harga(my_price: float, competitors: list) -> dict:
    """
    Menganalisa harga produk milik user dibandingkan dengan rata-rata pasar kompetitor.
    
    Args:
        my_price (float): Harga produk milik pengguna.
        competitors (list): Data mentah kompetitor hasil scraping.
        
    Returns:
        dict: Metrik statistik dan rekomendasi penetapan harga (pricing).
    """
    if not competitors: 
        return {}

    clean_comps, outlier_comps = filter_outlier_iqr(competitors)

    if outlier_comps:
        print(f"\n[INFO] Filter IQR: {len(outlier_comps)} produk outlier dibuang dari analisis:")
        for o in outlier_comps:
            name_trunc = o['nama_produk'][:55] + "..." if len(o['nama_produk']) > 55 else o['nama_produk']
            print(f"       -> Rp {o['harga']:,} | {name_trunc}")

    prices = [c["harga"] for c in clean_comps]
    if not prices: 
        prices = [c["harga"] for c in competitors if c["harga"] > 0]
        if not prices: return {}

    avg    = statistics.mean(prices)
    median = statistics.median(prices)
    q1     = statistics.quantiles(prices, n=4)[0] if len(prices) >= 4 else min(prices)
    q3     = statistics.quantiles(prices, n=4)[2] if len(prices) >= 4 else max(prices)
    
    pi       = my_price / avg
    diff     = my_price - avg
    diff_pct = (diff / avg) * 100
    gap_pct  = ((my_price - min(prices)) / min(prices)) * 100

    if pi < 0.97:    posisi, emoji = "SANGAT MURAH",    "🟢"
    elif pi < 1.00:  posisi, emoji = "MURAH",            "🟢"
    elif pi <= 1.03: posisi, emoji = "RATA-RATA PASAR", "🟡"
    else:            posisi, emoji = "MAHAL",            "🔴"

    iklan_ok = pi <= 1.03

    if pi < 0.97:
        rek, cvr, det = "✅ SANGAT DIREKOMENDASIKAN untuk iklan Shopee Ads", "Sangat kompetitif — CVR tinggi", "Harga sangat kompetitif. CVR akan tinggi karena buyer langsung tertarik."
    elif pi < 1.00:
        rek, cvr, det = "✅ DIREKOMENDASIKAN untuk iklan Shopee Ads", "Kompetitif — CVR aman", "Harga di bawah rata-rata. Aman diiklankan. Pastikan foto menarik."
    elif pi <= 1.03:
        rek, cvr, det = "⚠️ BOLEH iklan, tapi optimalkan dulu", "Netral — CVR masih bisa bersaing", "Harga rata-rata pasar. Tambahkan voucher/free ongkir agar CVR kompetitif."
    else:
        rek, cvr, det = "❌ TIDAK DIREKOMENDASIKAN iklan sekarang", "Kurang kompetitif — CVR akan rendah", f"Harga {abs(diff_pct):.1f}% di atas rata-rata. Turunkan ke Rp {int(avg*0.99):,} terlebih dahulu."

    return dict(
        avg=avg, median=median, min=min(prices), max=max(prices),
        q1=q1, q3=q3, price_index=pi, diff=diff, diff_pct=diff_pct,
        gap_pct=gap_pct, posisi=posisi, emoji=emoji, iklan_ok=iklan_ok,
        cvr=cvr, rek=rek, det=det, target=int(avg*0.99),
        n_clean=len(prices), n_outlier=len(outlier_comps), outlier_comps=outlier_comps
    )


def show_rich(product_name: str, my_price: float, comps: list, h: dict):
    """Menampilkan antarmuka laporan yang rapi di terminal menggunakan library Rich."""
    c = Console()
    c.print()
    c.print(Panel.fit(
        f"[bold]Shopee Price Research[/bold]\n[dim]{datetime.now().strftime('%d %b %Y  %H:%M')}[/dim]",
        border_style="blue"))
    c.print(f"\n[bold]Produk    :[/bold] {product_name}")
    c.print(f"[bold]Harga saya:[/bold] Rp {int(my_price):,}\n")

    outlier_text = ""
    if h['n_outlier'] > 0:
        formatted_outliers = [f"Rp {o['harga']:,} ({o['nama_produk'][:15]}..)" for o in h['outlier_comps']]
        if len(formatted_outliers) > 2:
            outlier_text = f"  [dim]({h['n_outlier']} dibuang: {', '.join(formatted_outliers[:2])}, dsb)[/dim]"
        else:
            outlier_text = f"  [dim]({h['n_outlier']} dibuang: {', '.join(formatted_outliers)})[/dim]"

    t = Table(title="Statistik Harga Kompetitor", box=box.ROUNDED)
    t.add_column("Metrik", style="cyan", min_width=24)
    t.add_column("Nilai",  style="white", justify="right")
    for k, v in [
        ("Jumlah data scrape",  f"{len(comps)} produk"),
        ("Data dianalisis",     f"[green]{h['n_clean']} produk[/green]{outlier_text}"),
        ("Harga minimum",       f"Rp {int(h['min']):,}"),
        ("Harga maksimum",      f"Rp {int(h['max']):,}"),
        ("Median",              f"Rp {int(h['median']):,}"),
        ("Rata-rata (bersih)",  f"Rp {int(h['avg']):,}"),
        ("Cluster Q1",          f"Rp {int(h['q1']):,}"),
        ("Cluster Q3",          f"Rp {int(h['q3']):,}"),
    ]:
        t.add_row(k, v)
    c.print(t); c.print()

    clr = "green" if h["iklan_ok"] else "red"
    p = Table(title="Analisa Posisi Harga", box=box.ROUNDED)
    p.add_column("Indikator", style="cyan", min_width=28)
    p.add_column("Nilai", justify="right")
    sign = "-" if h["diff"] < 0 else "+"
    for k, v in [
        ("Price Index",          f"{h['price_index']:.2f}"),
        ("Selisih vs rata-rata", f"{sign} Rp {abs(int(h['diff'])):,}"),
        ("Persentase",           f"{h['diff_pct']:+.2f}%"),
        ("Gap dari termurah",    f"{h['gap_pct']:+.1f}%"),
        ("Harga target",         f"Rp {h['target']:,}"),
        ("Posisi harga",         f"[bold {clr}]{h['emoji']}  {h['posisi']}[/bold {clr}]"),
    ]:
        p.add_row(k, v)
    c.print(p); c.print()

    c.print(Panel(
        f"[bold]{h['rek']}[/bold]\n\n{h['det']}\n\n[dim]CVR: {h['cvr']}[/dim]",
        title="[bold]Rekomendasi Iklan Shopee Ads[/bold]",
        border_style="green" if h["iklan_ok"] else "red"))
    c.print()

    sorted_c = sorted(comps, key=lambda x: x["total_terjual"], reverse=True)[:15]
    tbl = Table(title="Top 15 Kompetitor (terlaris)", box=box.SIMPLE_HEAVY)
    tbl.add_column("No",          style="dim",    width=4,  justify="right")
    tbl.add_column("Nama Produk", style="white",  min_width=32, max_width=45)
    tbl.add_column("Lokasi",      style="magenta")
    tbl.add_column("Harga",       style="yellow", justify="right")
    tbl.add_column("Terjual",     style="cyan",   justify="right")
    tbl.add_column("vs Saya",     justify="right")
    
    for i, x in enumerate(sorted_c, 1):
        vs = x["harga"] - my_price
        vc = "green" if vs >= 0 else "red"
        loc = x.get("lokasi", "-")
        
        # Logika Deteksi Toko Sendiri (Surabaya + Harga Sama)
        is_my_shop = (vs == 0 and "surabaya" in loc.lower())
        
        if len(loc) > 15: loc = loc[:12] + "..."
        
        vs_text = f"[{vc}]{'+'if vs>=0 else ''}Rp {vs:,}[/{vc}]"
        name_text = x['nama_produk']
        
        if is_my_shop:
            vs_text += "\n[cyan bold](Toko Kamu?)[/cyan bold]"
            name_text = f"[cyan bold]{x['nama_produk']}[/cyan bold]"
            
        tbl.add_row(str(i), name_text, loc, f"Rp {x['harga']:,}",
                    str(x["total_terjual"]), vs_text)
    c.print(tbl)
    c.print()


def show_plain(product_name: str, my_price: float, comps: list, h: dict):
    """Menampilkan laporan dalam format teks biasa jika library Rich tidak tersedia."""
    sep = "=" * 62
    print(f"\n{sep}\n  SHOPEE PRICE RESEARCH\n  {datetime.now().strftime('%d %b %Y  %H:%M')}\n{sep}")
    print(f"Produk     : {product_name}\nHarga saya : Rp {int(my_price):,}")
    print("-" * 62)
    print(f"Data scrape  : {len(comps)} produk")
    
    outlier_note = f"  ({h['n_outlier']} outlier dibuang)" if h['n_outlier'] > 0 else ""
    print(f"Dianalisis   : {h['n_clean']} produk{outlier_note}")
    
    for k, v in [("Minimum","min"),("Maksimum","max"),("Rata-rata (bersih)","avg"),
                 ("Median","median"),("Q1","q1"),("Q3","q3")]:
        print(f"{k:<20}: Rp {int(h[v]):,}")
    print("-" * 62)
    sign = "-" if h["diff"] < 0 else "+"
    print(f"Price Index  : {h['price_index']:.2f}")
    print(f"Selisih      : {sign} Rp {abs(int(h['diff'])):,}  ({h['diff_pct']:+.2f}%)")
    print(f"Harga target : Rp {h['target']:,}")
    print(f"\nPOSISI HARGA : {h['emoji']} {h['posisi']}\n{sep}")
    print(f"\nREKOMENDASI IKLAN:\n{h['rek']}\n\n{h['det']}\n\nCVR: {h['cvr']}\n{sep}")

    print("\n--- Top 15 Kompetitor (terlaris) ---")
    for i, x in enumerate(sorted(comps, key=lambda c: c["total_terjual"], reverse=True)[:15], 1):
        vs = x["harga"] - my_price
        loc = x.get("lokasi", "-")
        
        is_my_shop = (vs == 0 and "surabaya" in loc.lower())
        mark = " (🏪 KEMUNGKINAN TOKO KAMU)" if is_my_shop else ""
        
        print(f"\n{i:2}. {x['nama_produk'][:55]}{mark}")
        print(f"    Harga: Rp {x['harga']:,}  |  Terjual: {x['total_terjual']}  |  Lokasi: {loc}  |  vs Saya: {'+'if vs>=0 else''}Rp {vs:,}")
    print("\n" + sep)


def export_to_excel(all_results: list, filename: str):
    """
    Mengekspor seluruh hasil analisa (Summary, Top 15, Raw Data) ke dalam file Excel (.xlsx).
    
    Args:
        all_results (list): Daftar dictionary berisi keyword dan hasil analisa.
        filename (str): Nama output file excel.
    """
    if not HAS_OPENPYXL:
        print("\n[WARNING] Modul openpyxl tidak ditemukan, export excel dilewati.")
        return

    wb = Workbook()
    
    # Sheet 1: Summary Analisa
    ws_summary = wb.active
    ws_summary.title = "Summary Analisa"
    headers_summary = [
        "Keyword Produk", "Harga Saya", "Status Posisi", "Target Harga Optimal", 
        "Min", "Q1", "Median", "Rata-Rata (Avg)", "Q3", "Max",
        "Rekomendasi Utama", "Detail Analisa CVR",
        "Selisih (Rp)", "Persentase Gap", "Jml Data Dianalisa", "Jml Outlier", 
        "Detail Produk Outlier (Harga & Nama)"
    ]
    ws_summary.append(headers_summary)
    for cell in ws_summary["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sheet 2: Top 15 Terlaris
    ws_top = wb.create_sheet(title="Top 15 Terlaris")
    headers_top = ["Keyword Produk", "Peringkat", "Nama Produk Kompetitor", "Harga", "Total Terjual", "Lokasi", "Selisih vs Harga Saya", "Keterangan"]
    ws_top.append(headers_top)
    for cell in ws_top["1:1"]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Sheet 3: Semua Data Kompetitor
    ws_data = wb.create_sheet(title="Semua Data Kompetitor")
    headers_data = ["Keyword", "Nama Produk", "Harga", "Terjual", "Lokasi"]
    ws_data.append(headers_data)
    for cell in ws_data["1:1"]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for res in all_results:
        kw = res["keyword"]
        my_price = res["my_price"]
        h = res["h"]
        comps = res["comps"]
        
        if h:
            if h['n_outlier'] > 0:
                list_outliers_str = "\n".join([f"Rp {o['harga']:,} | {o['nama_produk']}" for o in h['outlier_comps']])
            else:
                list_outliers_str = "-"
                
            ws_summary.append([
                kw, my_price, f"{h['emoji']} {h['posisi']}", h['target'],
                h['min'], int(h['q1']), int(h['median']), int(h['avg']), int(h['q3']), h['max'],
                h['rek'], h['cvr'] + " - " + h['det'],
                int(h['diff']), f"{h['diff_pct']:.2f}%", h['n_clean'], h['n_outlier'], list_outliers_str
            ])
            ws_summary.cell(row=ws_summary.max_row, column=17).alignment = Alignment(wrap_text=True, vertical="top")
        else:
            ws_summary.append([kw, my_price] + ["Gagal Dianalisa"]*15)

        top_15 = sorted(comps, key=lambda x: x["total_terjual"], reverse=True)[:15]
        for i, c in enumerate(top_15, 1):
            selisih = c["harga"] - my_price
            keterangan = "Kemungkinan Toko Kamu" if (selisih == 0 and "surabaya" in c.get("lokasi", "").lower()) else ""
            
            ws_top.append([kw, i, c["nama_produk"], c["harga"], c["total_terjual"], c["lokasi"], selisih, keterangan])
            
        for c in comps:
            ws_data.append([kw, c["nama_produk"], c["harga"], c["total_terjual"], c["lokasi"]])

    # Auto-adjust column width
    for ws in [ws_summary, ws_top, ws_data]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if ws == ws_summary and col_letter == 'Q':
                    continue
                try:
                    if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                except: pass
            
            if ws == ws_summary and col_letter == 'Q':
                ws.column_dimensions[col_letter].width = 65 
            else:
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(filename)
    print(f"\n✅ Laporan komprehensif berhasil disimpan ke Excel: {filename}")


def main():
    """Fungsi utama yang menjalankan flow input, inisialisasi browser, scraping, dan export."""
    print("╔════════════════════════════════════════════════════════════╗")
    print("║ Shopee Price Analyzer (Bulk Scraping & Competitor Intel)   ║")
    print("╚════════════════════════════════════════════════════════════╝\n")

    print("Pilih Mode Input:")
    print("1. Single Keyword (Ketik manual)")
    print("2. Bulk Input (Dari file .txt)")
    
    mode = input("Masukkan pilihan (1/2): ").strip()
    items_to_scrape = []
    
    if mode == "1":
        kw = input("\nNama produk / keyword di Shopee:\n> ").strip()
        if not kw: return
        try:
            pr = float(input("Harga produk kamu (Rp):\n> ").strip().replace(".", "").replace(",", ""))
            items_to_scrape.append({"keyword": kw, "price": pr})
        except ValueError:
            print("Harga tidak valid."); return
            
    elif mode == "2":
        txt_path = input("\nMasukkan nama file .txt (contoh: produk.txt):\n> ").strip()
        if not os.path.exists(txt_path):
            print(f"[ERROR] File '{txt_path}' tidak ditemukan."); return
            
        with open(txt_path, "r", encoding="utf-8") as f:
            for line in f:
                if "|" in line:
                    parts = line.split("|")
                    try:
                        items_to_scrape.append({
                            "keyword": parts[0].strip(), 
                            "price": float(parts[1].strip().replace(".", "").replace(",", ""))
                        })
                    except ValueError: pass
        if not items_to_scrape: return
    else:
        return

    driver = setup_driver()
    all_results = []
    
    try:
        for idx, item in enumerate(items_to_scrape, 1):
            kw = item["keyword"]
            my_price = item["price"]
            
            print("\n" + "="*80)
            print(f"🚀 MEMPROSES ITEM {idx}/{len(items_to_scrape)}: {kw}")
            print("="*80)
            
            # Scrape data dari DOM
            comps = scrape_shopee(driver, kw, max_items=50, auto_mode=(mode=="2"))
            
            if not comps:
                print(f"[GAGAL] Tidak ada data terbaca untuk '{kw}'.")
                all_results.append({"keyword": kw, "my_price": my_price, "h": {}, "comps": []})
                continue
                
            # Analisa statistik
            h = analisa_harga(my_price, comps)
            
            if not h:
                all_results.append({"keyword": kw, "my_price": my_price, "h": {}, "comps": comps})
                continue

            # Simpan hasil untuk Excel
            all_results.append({
                "keyword": kw, 
                "my_price": my_price, 
                "h": h, 
                "comps": comps
            })
            
            # Print tampilan antarmuka
            if HAS_RICH:
                show_rich(kw, my_price, comps, h)
            else:
                show_plain(kw, my_price, comps, h)
            
            # Jeda antar pencarian saat bulk mode
            if mode == "2" and idx < len(items_to_scrape):
                time.sleep(2)
            
    finally:
        try: driver.quit()
        except: pass

    # Export Data Prompts
    if all_results:
        print("\n" + "="*80)
        pilihan = input("Simpan laporan keseluruhan ke Excel (.xlsx)? (y/n): ").strip().lower()
        if pilihan == 'y':
            fname = f"Shopee_Report_Analyzer_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            export_to_excel(all_results, fname)
        else:
            print("\n[INFO] Laporan tidak disimpan.")

if __name__ == "__main__":
    main()
